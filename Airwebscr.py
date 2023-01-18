import requests
from bs4 import BeautifulSoup
import openpyxl

# specify the url
url = 'https://www.example.com/flights'

# send a request to the website
response = requests.get(url)

# parse the HTML content of the website
soup = BeautifulSoup(response.content, 'html.parser')

# find all elements with the class 'ticket-price'
ticket_prices = soup.find_all(class_='ticket-price')

# open existing workbook
workbook = openpyxl.load_workbook("prices.xlsx")

# select the sheet to update
sheet = workbook["Sheet1"]

# get the current number of rows
current_row = sheet.max_row

# add the ticket prices to the sheet
for price in ticket_prices:
    sheet.cell(row=current_row + 1, column=1).value = price.text
    current_row += 1

# save the workbook
workbook.save("prices.xlsx")
